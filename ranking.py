import openpyxl
import csv
import pandas as pd
import os
import subprocess
import sys
import numpy as np
from pagerank2 import power_iteration as pr

def xlsxToCsv( wbFile ):
    #df = pd.read_excel( wbFile )#, sheetname='my_sheet_name')  # sheetname is optional
    #df.to_csv(wbFile.split(".")[0]+".html", index=False)  # index=False prevents pandas to write row index

    wb = openpyxl.load_workbook( wbFile )
    sh = wb.active
    with open(wbFile.split(".")[0]+".csv", 'w', newline="") as f:  # open('test.csv', 'w', newline="") for python 3
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    return wbFile.split(".")[0]+".csv"
    wb.close()


def csvToHTML( csvFile, name ): 
    # to read csv file
    a = pd.read_csv( csvFile )
     
    # to save as html file
    # named as "Table"
    # a.to_html( csvFile.split(".")[0]+".html" )
     
    # assign it to a
    # variable (string)
    table = a.to_html()
    table = table.split('\n', 1)[1]
    table = '<table id="{}" class="display cell-border">\n'.format( name )+table
    
    return table


def xlsxToHTML( wbFile, name ):
    csvFile = xlsxToCsv( wbFile )
    fileString = csvToHTML( csvFile, name )
    os.remove( csvFile )
    return fileString

def replaceInFile( destFile, startLine, endLine, payload ):
    skip = False
    with open( destFile, 'r') as f_in:
        with open(destFile+".tmp",'w') as f_out:
            for line_no, line in enumerate(f_in, 1):
                #print( line )
                if not skip:
                    f_out.write( line )
                if startLine in line:
                    #print( "HI" )
                    f_out.write( payload+"\n" )
                    skip = True
                if line.strip() == endLine:
                    f_out.write( line )
                    skip = False
    os.remove( destFile+".bak" )
    os.rename( destFile, destFile+".bak" )
    os.rename( destFile+".tmp", destFile )


def sendFileToPi( fileName, portFile ):
    """third line of credentials.txt in parent directory contains ssh port"""
   
    # after hours, used shlex to figure out what the hell Popen wants,
    # troubleshooting was annoying because of backslash
    # and spaces in windows directories
    # https://docs.python.org/3/library/subprocess.html#converting-argument-sequence
    path = os.path.abspath( fileName ).replace( "\\", "\\\\" )
    f = open(sys.path[0] + '/../'+portFile)
    f.readline()
    f.readline()
    port = f.readline().strip()
    dest = f.readline().strip()
    f.close()
   

    args = ['scp', '-r', '-P', port, path, dest]
    #print( args )
    p = subprocess.Popen( args )
    #print( p.returncode )
    p.wait()
    #sts = os.waitpid(p.pid, 0)

def updateBanDataWeb( banFile ):
    # update bans
    bantable = xlsxToHTML( banFile, "bans" )
    replaceInFile( "www/index.html", "<!STARTBANTABLE--><!--", "--><!ENDBANTABLE-->", bantable )
    sendFileToPi( "www/index.html", "credentials.txt" )


    
    
def updateGameDataWeb( sumFile, rankFile ):
    # oneliner
        #pd.read_excel('my_file', sheetname='my_sheet_name').to_csv('output_file_name', index=False)
        #wb = Workbook( wbFile )
        #wb.save( wbFile.split(".")[0]+".html" )
        #wb.close()



    # updateRankings
    buildRankingTable( sumFile, rankFile )

    ranktable = xlsxToHTML( rankFile, "rankings" )

    #print( ranktable )
    
    replaceInFile( "www/index.html", "<!STARTRANKTABLE-->", "<!ENDRANKTABLE-->", ranktable )

    

    # put game table in file
    # insert gameSummaries into html
    gametable = xlsxToHTML( sumFile, "games" )
    # print( table )
    replaceInFile( "www/index.html", "<!STARTGAMETABLE-->", "<!ENDGAMETABLE-->", gametable )


    

    # upload file to server
    sendFileToPi( "www/index.html", "credentials.txt" )
    #sendFileToPi( "www/jquery-3.6.0.js", "credentials.txt" )
    #print( xlsxToHTML( "gameSummaries.xlsx" ) )

def buildRankingTable( sumFile, rankFile ):
    # start here tomorrow. After ranking table is built with all the info needed for pagerank
    # implement changing the setup to allstars and hosting new rooms
    # and closing rooms
    wb = openpyxl.load_workbook( sumFile )
    
    games = wb.active

    # create dictionary with keys usernames and values of the following format
    # {"townWinsOn": {"beanstalk":2, "sample":1}, "mafiaWinsOn": {"sample":5}, "townLossTo": {}, "mafiaLossTo": {"sample":1,"beanstalk":3}}

    data = {}

    row = 2
    while not games['A'+str(row)].value is None:
        winTeam = games['C'+str(row)].value

        if winTeam == "None":
            row += 1
            continue
        
        if winTeam == "Cult":
            winTeam = "Town"

        winners = games['E'+str(row)].value
        #print( type( winners ), winners )
        
        winners = winners.replace(",","").split()
        losers = games['G'+str(row)].value
        
        losers = losers.replace(",","").split()
        #print( winTeam, winners, losers )
        

        for winner in winners:
            if winner == "-":
                break
            if not winner in data:
                data[winner] = {"pr": "-","mpr":"-","rank":"-","mrank":"-","winsAsTown": 0, "winsAsMafia": 0, "lossesAsTown": 0, "lossesAsMafia": 0, "townWinsOn": {}, "mafiaWinsOn": {}, "townLossTo": {}, "mafiaLossTo": {}}
            if winTeam == "Town":
                data[winner]["winsAsTown"]+=1
            if winTeam == "Mafia":
                data[winner]["winsAsMafia"]+=1
        for loser in losers:
            if loser == "-":
                break
            if not loser in data:
                data[loser] = {"pr": "-", "mpr":"-","rank":"-","mrank":"-","winsAsTown": 0, "winsAsMafia": 0, "lossesAsTown": 0, "lossesAsMafia": 0, "townWinsOn": {}, "mafiaWinsOn": {}, "townLossTo": {}, "mafiaLossTo": {}}
            if winTeam == "Town":
                data[loser]["lossesAsMafia"]+=1
            if winTeam == "Mafia":
                data[loser]["lossesAsTown"]+=1


        if winners == ["-"]:
            row += 1
            continue
        if losers == ["-"]:
            row += 1
            continue
            
        # possibly add edges between town members to see if this rewards a town win more
        townCoopBoost = True
        
        
        for winner in winners:
            for loser in losers:
                #if not winner in data:
                #        data[winner] = {"townWinsOn": {}, "mafiaWinsOn": {}, "townLossTo": {}, "mafiaLossTo": {}}
                #if not loser in data:
                #        data[loser] = {"townWinsOn": {}, "mafiaWinsOn": {}, "townLossTo": {}, "mafiaLossTo": {}}
                if winTeam == "Town":
                    # edges mafia to town
                    try:
                        data[winner]["townWinsOn"][loser] += 1
                    except KeyError as msg:
                        data[winner]["townWinsOn"][loser] = 1
                    try:
                        data[loser]["mafiaLossTo"][winner] += 1
                    except KeyError as msg:
                        data[loser]["mafiaLossTo"][winner] = 1
                    # possibly add edges between town members to see if this rewards a town win more
                    if townCoopBoost:
                        pass                        
                if winTeam == "Mafia":
                    # edges from town to mafia
                    try:
                        data[winner]["mafiaWinsOn"][loser] += 1
                    except KeyError as msg:
                        data[winner]["mafiaWinsOn"][loser] = 1
                    try:
                        data[loser]["townLossTo"][winner] += 1
                    except KeyError as msg:
                        data[loser]["townLossTo"][winner] = 1

        row += 1

    wb.close()

    # compute pagerank

    prDict = myPageRank( data )
    for key in prDict:
        data[key]["pr"]=prDict[key]
    

    # build ranking table

    wb = openpyxl.load_workbook( rankFile )
    
    rankings = wb.active
    

    numGames = row - 2

    # print( "data ", data )


    # calculating modified PageRank which weights how many games
    # each player has played as compared to the average. more games
    # played is a boost in rank, regardless if win or loss
    wFactor = 0 # how much to weight this?
    gamesPlayed = {}
    gameTotal = 0
    for player in data:
        tWins = data[player]["winsAsTown"]
        mWins = data[player]["winsAsMafia"]
        tLoss = data[player]["lossesAsTown"]
        mLoss = data[player]["lossesAsMafia"]
        games = tWins+mWins+tLoss+mLoss
        gamesPlayed[ player ] = games
        gameTotal += games
    avgGamesPerPlayer = gameTotal/len( data )
    
    print( "Average games per player:", avgGamesPerPlayer )
    mprDict = {}
    for player in data:
        mprDict[player] = (1 - wFactor) * prDict[player]+ wFactor * gamesPlayed[player]/gameTotal
        data[player]["mpr"]=mprDict[player]

    # players below the threshold are unranked:
    for player in data:
        if gamesPlayed[ player ] < avgGamesPerPlayer:
            data[player]["pr"] = "-"
            prDict[player] = 0
            data[player]["mpr"] = "-"
            mprDict[player] = 0
    
    # calculate rank based on this:
    sortedMRank  = sorted(mprDict.items(), key = lambda kv:(kv[1], kv[0]), reverse = True)
    sortedRank  = sorted(prDict.items(), key = lambda kv:(kv[1], kv[0]), reverse = True)
    #print( sortedRank )
    rankDict = {}
    for i in range( len( sortedMRank ) ):
        data[ sortedMRank[i][0] ]["mrank"] = i+1
    for i in range( len( sortedRank ) ):
        data[ sortedRank[i][0] ]["rank"] = i+1

        
        

    
    
    checksum = 0
    row = 2
    for player in data:
        tWins = data[player]["winsAsTown"]
        mWins = data[player]["winsAsMafia"]
        tLoss = data[player]["lossesAsTown"]
        mLoss = data[player]["lossesAsMafia"]

        if tWins+tLoss != 0:
            tWinPercent = tWins/(tWins+tLoss)
            tWinPercent = "{:.2f}%".format( tWinPercent*100 )
        else:
            tWinPercent = "-"

        if mWins+mLoss != 0:
            mWinPercent = mWins/(mWins+mLoss)
            mWinPercent = "{:.2f}%".format( mWinPercent*100 )
        else:
            mWinPercent = "-"


        winPercent = (tWins+mWins)/(tWins+mWins+tLoss+mLoss)


        if not type( data[player]["pr"] ) == str:
            prank = "{:.4f}%".format( data[player]["pr"]*100 )
            rank = data[player]["rank"]
        else:
            prank = "-"
            rank = 9999999
        if not type( data[player]["mpr"] ) == str:
            mprank = "{:.4f}%".format( data[player]["mpr"]*100 )
            mrank = data[player]["mrank"]
        else:
            mprank = "-"
            mrank = 9999999
            
        #checksum += data[player]["mpr"] useless to have this if setting things to 0       
        
        
        rankings['A'+str(row)] = player # username
        rankings['B'+str(row)] = str( data[player] ) # raw data

        rankings['C'+str(row)] = mprank # Modified PageRank
        rankings['D'+str(row)] = mrank # rank based on modified PageRank
        

        
        rankings['E'+str(row)] = prank # PageRank
        rankings['F'+str(row)] = rank # rank based on pure PageRank
        

        rankings['G'+str(row)] = tWins+mWins+tLoss+mLoss # total games played
        rankings['H'+str(row)] = str( tWins+mWins )+":"+str( tLoss+mLoss ) #total win:loss ratio
        rankings['I'+str(row)] = tWins+mWins # raw wins
        rankings['J'+str(row)] = "{:.2f}%".format( winPercent*100 ) # raw win percentage

        rankings['K'+str(row)] = tWins+tLoss # games as town
        rankings['L'+str(row)] = str( tWins )+":"+str( tLoss )#town win:loss ratio
        rankings['M'+str(row)] = tWinPercent
             
        
        rankings['N'+str(row)] = mWins+mLoss # games as mafia  
        rankings['O'+str(row)] = str( mWins )+":"+str( mLoss )#mafia win:loss ratio
        rankings['P'+str(row)] = mWinPercent # win percentage as mafia
        
        
        
        
        
        
        
        
        
        
        #print( key, data[key] )
        row += 1

    #print( "Checksum: ", checksum )
    wb.save( rankFile )

    wb.close()
                        
                            
                            

    #for row in range( 2, 


    
def myPageRank( myDict ):
    indexToKey = {}
    keyToIndex = {}
    i=0
    for key in myDict:
        keyToIndex[key] = i
        indexToKey[i]=key
        i+=1

    mat = []
    for key in myDict: # building each row
        row = [0]*len( myDict )
        for lostAgainst in myDict[ key ]["townLossTo"]:
            row[ keyToIndex[lostAgainst] ] += myDict[ key ]["townLossTo"][ lostAgainst ]
        for lostAgainst in myDict[ key ]["mafiaLossTo"]:
            row[ keyToIndex[lostAgainst] ] += myDict[ key ]["mafiaLossTo"][ lostAgainst ]

        #for j in range( len( myDict ) ):
        #    if myDict[key]
            


        mat.append( row )


    # I have no idea if this is reasonable but I like the
    # idea of allowing nodes to point back to themselves-
    # seems to make the distribution more extreme
    # Tweaking this to anything positive has this effect
    # Larger values bring things closer to uniform
    #diagAdd = max( [ max( row ) for row in mat ] )
    totalSum = 0
    for row in mat:
        totalSum += sum( row )
    diagAdd = totalSum
    print( diagAdd )
    
    for i in range( len( mat ) ):        
        mat[i][i] += diagAdd

    mat = np.array( mat )
    ranks = pr( mat, rsp=0.15, epsilon=0.00001, max_iterations=1000 )
    #print( ranks )
    prDict = {}
    for i in range( len( ranks ) ):
        print( ranks[i] )
        prDict[ indexToKey[i] ] = ranks[i]

    print( prDict )
    print( keyToIndex )

    print( mat )
    return prDict
    #print( ranks )
        
    
    #print( myDict )
    #print( keyToIndex )

if __name__=='__main__':
    updateGameDataWeb( "gameSummaries.xlsx", "rankings.xlsx" )
    updateBanDataWeb( "bans.xlsx" )
    
    
