from websockets import connect, WebSocketClientProtocol
import requests as rq
import sys
from json import loads, dumps
from time import sleep
import asyncio
import traceback
import openpyxl
from datetime import datetime
import ranking

DEBUG = False

# tomorrow, try this for creating a sortable table:
# https://stackoverflow.com/questions/10683712/html-table-sort
# datatables looks good

#class Game:
#    def __init__( self ):
#        self.phase = "Pregrame" #Phase is one of Pregrame, Ingame, Postgame
        
# convert curl request (can be found using devtools) to python request: https://curl.trillworks.com/#python

# check mafiascum/mafiauniverse to see what other stats might be worth recording

class Room:
    def __init__( self, token, roomName, startedBy, listed, referer ):
        #cookies = token

        data = {'name': roomName, 'unlisted': not listed }
        headers = {}
        if referer:
            headers['Referer']=referer

        #print( referer )

        #print( "hi" )
        # will crash below if on vpn (banned)
        self.roomId = loads( rq.post('https://mafia.gg/api/rooms', cookies=token, data=data, headers=headers).content )['id']
        #print( "bye" )
        
        self.startedBy = startedBy
        self.phase = "pre" #Phase is one of pre, in, post
        #self.lastsid = 0 # sid increases for each game event
        #print( "hi" )
        #sleep( 30 )
        #print( self.roomId )
        #roomID = input( "Enter room id: " )
        #self.roomName = roomName
        #self.listed = listed
        #roomOptions = {'name': roomName, 'unlisted': not listed}
        #tmp = rq.post('https://mafia.gg/api/rooms', json=roomOptions )#, cookies=token )#, cookies=token).content
        #print( type( tmp ) )
        #print( "hi" )
        #room = loads(rq.post('https://mafia.gg/api/rooms', json=roomOptions, cookies=token).content)['id']
        self.engine = loads(rq.get('https://mafia.gg/api/rooms/{}'.format(self.roomId), cookies=token).content)
        #print( type( self.engine ) )
        #self.ws = None
        #self.info = None
        #self.users= None
        #self.ws = self.createWebSocket()
        #print( type( self.ws ) )
        #self.cachedInfo = self.updateInfo()
        self.gameEvents = {"pre":[],"in":[],"post":[]} # keys pre, in, post
        self.curInfo()
        self.startInfo = None
        
        
        #self.info = self.curInfo()
        
        #async def initAsync(token, roomName, startedBy, listed ):
        #    self.ws = await connect( self.engine['engineUrl'], ssl=True) # web socket for room?
            
        #    output = dumps({'type':'clientHandshake', 'userId':self.startedBy, 'roomId':self.roomId, 'auth': self.engine['auth']}) #

        #    await self.ws.send( output)
            
        #    self.info = loads( await self.ws.recv() )
            
            #self.users = self.info['users']
            

            
            #self.settings = loads(rq.get('https://mafia.gg/api/rooms/{}'.format(roomId), cookies=cookies).content)

            #print( "self.settings=", self.settings )
                

            #print( "self.ws=", self.ws )
            #output = dumps({'type':'clientHandshake', 'userId':userId, 'roomId':roomID, 'auth': self.settings['auth']})
            #self.ws.send(output)
            #print( "output: ", self.output )
            #self.info = loads(  self.ws.recv())
            #print( "info: ", self.info )
            #self.users = self.info['users']
            #print( "users: " , self.users )
            #room = loads(rq.post('https://mafia.gg/api/rooms', json=options, cookies=currentcookie).content)['id']
            #
            #self.sendPacket({'type': 'newGame', 'roomId': 234234})
            #make = mafiaConnectionold(room, make.userId, currentcookie)
            #pass
        #asyncio.get_event_loop().run_until_complete( initAsync( token, roomName, startedBy, listed ))
        #self.cachedInfo = self.updateInfo()
        #print( "websocket:", self.ws)
        
        #print( "Room info:", self.info )

    #def createWebSocket( self ):
    #    async def asyncCreateWebSocket():
    #        ws = await connect( self.engine['engineUrl'], ssl=True) # web socket for room?
    #    return asyncio.get_event_loop().run_until_complete( asyncCreateWebSocket() )
    """#self.default_options = {'dayLength': 3, 'dayStart': 'off',
               'deadlockPreventionLimit': '-1', 'deck': '-1',
               'disableVoteLock': False, 'hideSetup': False,
               'hostRoleSelection': False, 'majorityRule': '51',
               'mustVote': False, 'nightLength': 1, 'noNightTalk': False,
               'revealSetting': 'allReveal', 'roles': {},
               'name': 'TESTING', 'scaleTimer': True, 'twoKp': '0',
               'type': 'options', 'unlisted': False}"""

    def curInfo( self ):
        async def asyncUpdateInfo():
            self.ws = await connect( self.engine['engineUrl'], ssl=True) # web socket for room?
            output = dumps({'type':'clientHandshake', 'userId':self.startedBy, 'roomId':self.roomId, 'auth': self.engine['auth']})
            await self.ws.send( output)            
            recv = loads( await self.ws.recv() )
            return recv
        self.info = asyncio.get_event_loop().run_until_complete( asyncUpdateInfo() )
        #self.gameEvents[ self.phase ].append( self.info )
        return self.info

    def lastSettings( self ):
        options = {}
        #maxstamp = 0
        for i in range( len( self.info['events'] ) - 1, 0, -1 ):
            if self.info['events'][i]['type']=='options':
                options['settings']= self.info['events'][i]
                break
        #options['options']= self.info['events']
        options['users']=self.info['users']
        options['possibleUserIds'] = self.info['possibleUserIds']
        options['timestamp']=self.info['timestamp']
        options['sid']=self.info['sid']
        return options           

    def sendPacket( self, packet ):       
        async def asyncSendPacket( packet ):
            await self.ws.send( dumps( packet ) )
            recv = loads(await self.ws.recv())
            return recv
        tmp = asyncio.get_event_loop().run_until_complete( asyncSendPacket( packet ) )
        #self.gameEvents[ self.phase ].append( tmp )
        return tmp

    def optionsPacket( self, optionsPacket ):
        if optionsPacket["type"]=="options":
            sendPacket( optionsPacket )
    
    def getNextEvent( self ):
        async def asyncGet():
            recv = loads(await self.ws.recv())
            
            return recv
        tmp = asyncio.get_event_loop().run_until_complete( asyncGet() )
        #self.gameEvents[ self.phase ].append( tmp )
        return tmp
    
    def talk( self, msg ):
        #async def asyncSendMsg( message ):
        #    await self.ws.send(dumps({'type': 'chat', 'message': message}))
        #    recv = loads( await self.ws.recv() )
        #    self.info['events'].append( recv )
        #    return recv
        #return asyncio.get_event_loop().run_until_complete( asyncSendMsg( message ) )
        self.sendPacket( {'type': 'chat', 'message': msg} )

    def afkCheck( self ):
        # seems to hang if no users are present
        self.sendPacket( {'type': 'forceSpectate'} )

    def startGame( self ):
        #self.curInfo()
        return self.sendPacket({'type': 'startGame'})        

    def setup( self, code ):
        #{'roles': self.convertSetup(code)}
        def convertSetup(setup):
            return dict(map(lambda x:str.split(x, 'a'), str.split(setup, 'b')))
        converted = convertSetup( code )
        #print( converted )
        return( self.sendPacket( {'type': 'options', 'roles': converted }) )

    def setupAllStars( self, num ):
        payload = {'dayLength': 11, 'dayStart': 'mafiaNKn1',
               'deadlockPreventionLimit': '-1',
               'disableVoteLock': True, 'hideSetup': False,
               'hostRoleSelection': False, 'majorityRule': '51',
               'mustVote': False, 'nightLength': 3, 'noNightTalk': False,
               'revealSetting': 'noReveal', 'roles': {"31": 1,"34": 1,"38": 1,"57": 7,"75": 3},
               'roomName': '⭐Ranked All Stars (Game {})⭐ [mafiastats.cf]'.format( num ), 'scaleTimer': False, 'twoKp': '3',
               'type': 'options'}
        return self.sendPacket( payload )

    def setupCoffee( self ):
        payload = {'dayLength': 1, 'dayStart': 'on',
               'deadlockPreventionLimit': '-1',
               'disableVoteLock': True, 'hideSetup': False,
               'hostRoleSelection': False, 'majorityRule': '51',
               'mustVote': False, 'nightLength': 1, 'noNightTalk': False,
               'revealSetting': 'allReveal', 'roles': {"93": 1,"75": 1},
               'roomName': 'Debug 1', 'scaleTimer': True, 'twoKp': '0',
               'type': 'options', 'unlisted': True}
        return self.sendPacket( payload )

    def setup2p( self ):
        payload = {'dayLength': 1, 'dayStart': 'on',
               'deadlockPreventionLimit': '-1',
               'disableVoteLock': True, 'hideSetup': False,
               'hostRoleSelection': False, 'majorityRule': '51',
               'mustVote': False, 'nightLength': 1, 'noNightTalk': False,
               'revealSetting': 'allReveal', 'roles': {"57": 1,"75": 1},
               'roomName': 'Debug 2', 'scaleTimer': True, 'twoKp': '0',
               'type': 'options', 'unlisted': True}
        return self.sendPacket( payload )

    def emptySetup( self ):
        payload = {'roles': {},'type': 'options', 'unlisted': True}
        return self.sendPacket( payload )

    #def RunEmpty( self ):
    #    tmp1 = self.empty()
    #    tmp2 = self.
        

    def timers( self, day, night ):
        return( self.sendPacket( {'type': 'options', 'dayLength': day, 'nightLength': night }) )

    def disableVoteLock( self, flag ):
        """disableVoteLock is bool"""
        return( self.sendPacket( {'type': 'options', 'disableVoteLock': flag }) )

    def scaleTimer( self, flag ):
        return( self.sendPacket( {'type': 'options', 'scaleTimer': flag } ) )

    def list( self ):
        return( self.sendPacket( {'type': 'options', 'unlisted': False }) )

    def unlist( self ):
        return( self.sendPacket( {'type': 'options', 'unlisted': True }) )

    def becomePlayer( self ):
        self.sendPacket({'type': 'presence', 'isPlayer': True})
        
    def becomeSpectator( self ):
        self.sendPacket({'type': 'presence', 'isPlayer': False})        

    def unameFromId( self ):
        pass

    def addRole( self, amt, name ):
        pass

    def remRole( self, amt, name ):
        pass

    def kick( self, sesh, user ):
        #loads( rq.post('https://mafia.gg/api/rooms', cookies=token, data=data, headers=headers).content )['id']
        # needs fixed
        toKick = ""
        for item in self.getUsers():
            if sesh.usernameFromId( item['userId'] ) == user:
                toKick = item['userId']
        if toKick == "":
            return -1

        print( toKick )
        tmp = rq.post('https://mafia.gg/api/rooms/{}/kick'.format(self.roomId), cookies=sesh.token, json={'userId': toKick} ).content
        #tmp = loads( rq.post('https://mafia.gg/api/rooms/{}/kick'.format(self.roomId), cookies=sesh.token, json={'userId': toKick} ).content )
        return tmp
        
            
        #async def asyncKick( user ):
        #    return await loads( rq.post('https://mafia.gg/api/rooms/{}/kick'.format(self.roomId), cookies=sesh.token, json={'userId': toKick} ) )
        #tmp = asyncio.get_event_loop().run_until_complete( asyncKick( str( toKick ) ) )
        #return tmp        

    def url( self ):
        return "https://mafia.gg/game/"+self.roomId
    
    def getStartEventIndex( self ):
        for i in range( 1, len( self.gameEvents[ "in" ] ) ):
            try:
                if self.gameEvents[ "in" ][i]["type"] == "startGame":
                    return i
            except:
                traceback.print_exc()
                continue
        return -1

    def getUsers( self ):
        return self.curInfo()["users"]

    def getSettings( self ):
        self.curInfo()
        return lastSettings()           

    def writeGameToFile( self ):
        f = open(sys.path[0] + '/../games/'+self.roomId+".txt", "w", encoding='utf-8')
        #print( "pre:", self.gameEvents[ "pre" ] )
        #print( "in:", self.gameEvents[ "in" ] )
        #print( "post:", self.gameEvents[ "post" ] )
        f.write( str( self.gameEvents[ "in" ][0] )+"\n" )
        startIndex = self.getStartEventIndex()
        for i in range( startIndex, len( self.gameEvents[ "in" ] )):
            f.write( str( self.gameEvents[ "in" ][i] )+"\n" )
        f.close()
        return
        
        #startFlag = False
        #for i in range( 1, len( self.gameEvents[ "in" ] ) ):
        #    try:
        #        if self.gameEvents[ "in" ][i]["type"] != "startGame" and  not startFlag:
        #            continue
        #    except:
        #        traceback.print_exc()
        #        continue
        #    if not startFlag:
        #        self.startInfo = self.gameEvents[ "in" ][i]
        #    startFlag = True
        #    f.write( str( self.gameEvents[ "in" ][i] )+"\n" )
        #f.close()

    def writeGameSummary( self, sesh, summaryFile ):
        sumFile = summaryFile
        wb = openpyxl.load_workbook( sumFile )
        w = wb.active
        w['A1'] = 'RoomId'
        w['B1'] = 'Finished at' 
        w['C1'] = 'Winning alignment'
        w['D1'] = 'Winning team userIds'
        w['E1'] = 'Winning team'
        w['F1'] = 'Losing team userIds'
        w['G1'] = 'Losing team'
        row = 2
        while not w['A'+str(row)].value is None:
            row += 1

        winningTeams = self.gameEvents[ "in" ][-3]['message'].split(":")[1].replace( ",", "" ).split()
        players = self.gameEvents[ "in" ][ self.getStartEventIndex() ]['players']
        winnerString = self.gameEvents[ "in" ][-2]['message'].split(":")[1]
        loseIds = []
        winIds = []
        userToPlayer = self.gameEvents[ "in" ][-5]['users']
        #dt_object = datetime.fromtimestamp(timestamp)

        dt = str( datetime.fromtimestamp( self.gameEvents[ "in" ][-1]['timestamp'] ) )+" CST"
        print( dt )
        #print( players )
        #print( winnerString )
        #print( userToPlayer )
        #print( loseIds )
        #print( winIds )
        for player in players:
            if player['name'] in winnerString:
                for user in userToPlayer:
                    if userToPlayer[ user ] == player['playerId']:
                        winIds.append( user )
                        break
            else:
                for user in userToPlayer:
                    if userToPlayer[ user ] == player['playerId']:
                        loseIds.append( user )
                        break

        winners = [ sesh.usernameFromId( item ) for item in winIds ]
        losers = [ sesh.usernameFromId( item ) for item in loseIds ]
        if winners == []:
            winners = ['-']
        if winIds == []:
            winIds = ['-']
        if losers == []:
            losers = ['-']
        if loseIds == []:
            loseIds = ['-']


        print( winningTeams )
        print( players )
        print( loseIds )
        print( losers )
        print( winIds )
        print( winners )

        w['A'+str(row)] = self.roomId
        w['B'+str(row)] = dt
        winTeams = str( winningTeams )[1:-1].replace("'","")#.replace(" ","")
        w['C'+str(row)] = winTeams
        w['D'+str(row)] = str( winIds )[1:-1].replace("'","")#.replace(" ","")
        winners = str( winners )[1:-1].replace("'","")#.replace(" ","")
        w['E'+str(row)] = winners
        w['F'+str(row)] = str( loseIds )[1:-1].replace("'","")#.replace(" ","")
        w['G'+str(row)] = str( losers )[1:-1].replace("'","")#.replace(" ","")

        while True:
            try:
                wb.save( sumFile )
                break
            except PermissionError as msg:
                traceback.print_exc()
                self.talk( "Check terminal.")
                input( "Close the game summary spreadsheet to continue. <Enter> to try again." )
        wb.close()
        #return winTeams.replace(",",", "), winners.replace(",",", ")
        return winTeams, winners

class Session:

    commandChar = "\\"

    roomName = "temp"
    sumFile = "gameSummaries.xlsx"
    rankingFile = "rankings.xlsx"
    banFile = "bans.xlsx"
    #welcomeMsg1 = "⭐Ranked All Stars⭐"
    welcomeMsg2 = "Welcome! Rules and public rankings can be found at https://mafiastats.cf."#Type {}help for a list of available commands.".format( commandChar ) 
    afkMsg = "AFK check! Click 'Become a player' to rejoin."
    delay = 3
    startMsg = "Starting game in {} seconds...".format( delay )
    disclaimer = "DISCLAIMER: By participating in this game, you consent to your in-game logs being collected and to your username and win/loss data appearing in public rankings. You also consent to the use of your data to develop mafia AI. Finally, by playing this game, you agree to play to your win condition to the best of your ability, and to be kind and respectful to other players. Throwers/cheaters/AFKers/namecallers will be banned. Contact beanstalk#5555 on Discord for ban appeals."
    ratingStr = "Public ratings updated at https://mafiastats.cf."
    def __init__( self, credFile, dictFile ):
        """Begins a session using the credentials in credFile. Assumes
        the first line contains the username and the next line contains the
        password. Expects to find file in parent directory. Expects to find
        dictionaries of role info, etc, in file in current directory."""
        f = open(sys.path[0] + '/../'+credFile)
        self.authusers = f.readline().strip().split( "," )
        username = self.authusers[0]
        password = f.readline().strip()
        f.close()
        tmp = rq.post('https://mafia.gg/api/user-session', json={'login': username, 'password': password})
        #print( "hi" )
        
        self.userData = loads(tmp.text)
        self.token = tmp.cookies.get_dict()#['userSessionToken']
        self.roles = self.getDictionaries( dictFile )
        self.bans = self.readBans() #self.userData['hostBannedUsernames']
        self.nextRoomNum = self.nextRoomNumberFromFile()
        self.keepGoing = True
        self.recording = True
        
        self.rooms = [] # list of hosted rooms under this session
        del tmp

    def nextRoomNumberFromFile( self ):
        wb = openpyxl.load_workbook( self.sumFile )
        w = wb.active
        row = 2
        while not w['A'+str(row)].value is None:
            row += 1
        return row - 1
        wb.close()
        

    def readBans( self ):
        curBans = self.userData['hostBannedUsernames']
        toReturn = {}
        wb = openpyxl.load_workbook( self.banFile )
        w = wb.active
        row = 2
        while not w['A'+str(row)].value is None:
            for user in curBans:
                if w['A'+str(row)].value == user:
                    toReturn[ user ] = w['B'+str(row)].value
                    break
                    #break
                #else:
                #    continue
                #    toReturn[ user ] = "-"
                #    break
            row += 1

        wb.close()
        
        
        return toReturn

    def writeBans( self ):
        wb = openpyxl.load_workbook( self.banFile )
        w = wb.active
        row = 2
        for user in self.bans:
            w['A'+str(row)] = user
            w['B'+str(row)] = self.bans[ user ]
            row += 1
        while not w['A'+str(row)].value is None:
            w['A'+str(row)].value = None
            w['B'+str(row)].value = None
            row += 1
        wb.save( self.banFile )
        wb.close()
        
            
        

    def hostNewRoom( self, listed = False, referer = None  ):
        newRoom = Room( self.token, self.roomName, self.userData['id'], listed, referer )
        print( "Created room at:", newRoom.url() )
        newRoom.talk( "⭐Ranked All Stars (Game {})⭐".format( self.nextRoomNum ) )
        newRoom.talk( self.welcomeMsg2 )
        newRoom.talk( self.disclaimer )
        self.rooms.append( newRoom )
        newRoom.setupAllStars( self.nextRoomNum )
        self.nextRoomNum += 1
        return newRoom
        
    def __str__( self ):
        return "userData = " + str( self.userData )+ "\n" + "token = " + str( self.token ) + "\nActive rooms: " + str( self.rooms )

    def close( self ):
        pass

    def userInfo( self, userId ):
        return loads( rq.get( 'https://mafia.gg/api/users/{}'.format( userId ) ).content)[0]

    def usernameFromId( self, userId ):
        return self.userInfo( userId )['username']

    def roomInfo( self, roomId ):
        return loads( rq.get( 'https://mafia.gg/api/rooms/{}'.format( roomId ), cookies=self.token ).content)

    def getDictionaries( self, dictFile ):
        "Expects to find file in current directory containing global dictionaries."
        f = open( sys.path[0] + '/./'+dictFile )
        roles = loads( f.readline() )
        f.close()
        return roles

    def bansToWeb( self ):
        while True:
            try:
                self.writeBans()
                break
            except PermissionError as msg:
                traceback.print_exc()
                room.talk( "Check terminal.")
                input( "Close the bans spreadsheet to continue. <Enter> to try again." )
        try:
            ranking.updateBanDataWeb( self.banFile )
        except:
            traceback.print_exc()

    def blacklist( self, user, reason ):
        #print( "before: ", self.userData['hostBannedUsernames'] )
        if user not in self.bans:
            #headers={'content-type': 'application/json'}
            self.bans[ user ] = reason            
            data = {'hostBannedUsernames': list( self.bans.keys() ) }
            print( "data=",data )
            print( self.token )
            tmp = loads( rq.patch('https://mafia.gg/api/user', cookies=self.token, json=data).content )
            #tmp2 = loads( rq.get('https://mafia.gg/api/user', cookies=self.token, json=data, headers=headers).content )
                

            #print( "after: ", self.userData['hostBannedUsernames'] )
            self.bansToWeb()
            return tmp
        return "User already banned."


    def unblacklist( self, user ):
        #print( "before: ", self.userData['hostBannedUsernames'] )
        if user in self.bans:
            #headers={'content-type': 'application/json'}
            del self.bans[ user ]
            #self.bans.remove( user )            
            data = {'hostBannedUsernames': list( self.bans.keys() ) }
            print( "data=",data )
            print( self.token )
            tmp = loads( rq.patch('https://mafia.gg/api/user', cookies=self.token, json=data).content )
            #tmp2 = loads( rq.get('https://mafia.gg/api/user', cookies=self.token, json=data, headers=headers).content )
                

            #print( "after: ", self.userData['hostBannedUsernames'] )
            self.bansToWeb()
            return tmp
        return "User already unbanned."

    #def getbans( self ):
    #    return self.bans

    def clearbans( self ):
        self.bans = {}
        data = {'hostBannedUsernames': list( self.bans.keys() ) }
        tmp = loads( rq.patch('https://mafia.gg/api/user', cookies=self.token, json=data ).content )
        self.bansToWeb()
        return tmp
        

        # needs to be fixed
        #async def blacklist(self, user):
        #    this = self.user
        #    return await self.session.patch('https://mafia.gg/api/user', json=
        #        {'email': this.email, 'hostBannedUsernames': [*this.hostBannedUsernames, user.username],
        #         'password': '', 'passwordConfirmation': '', 'patreonCode': None, 'timeFormat': this.timeFormat})


    #def bansToFile( self ):
    #    pass
        

    def processPacket( self, room, update ):
        
        # check if endgame packet,
        # and write to file if endgame
        try:    
            if update[ "type" ] == 'endGame':       
                for i in range( 4 ):
                    update = room.getNextEvent()
                    room.gameEvents[ room.phase ].append( update )
                room.phase = "post"
                room.gameEvents[ room.phase ].append( update )
                
                
                
                    

                #room.talk( "Check terminal." ) 
                #toFile = input( "Write  game to file? (y/n) " )
                
                if self.recording:
                    room.writeGameToFile()
                    winTeams, winUsers = room.writeGameSummary( self, self.sumFile )

                
                
                                
                    room.talk( "Congrats to the members of " + str( winTeams ) + ": "+ str( winUsers ) + "! Updating public rankings..."  )

                    while True:
                        try:
                            ranking.updateGameDataWeb( self.sumFile, self.rankingFile )
                            break
                        except PermissionError as msg:
                            traceback.print_exc()
                            room.talk( "Check terminal.")
                            input( "Close the rankings spreadsheet to continue. <Enter> to try again." )
                            
                    room.talk( self.ratingStr )
                else:
                    room.talk( "The data from this game will not be recorded." )
                    self.nextRoomNum -= 1
                    
        except:
            traceback.print_exc()
                
            
        #start here tomorrow
        #print( update )
        user, command = "", ""
        try:
            if update[ "type" ] == "chat" and update[ "message"][0] == self.commandChar and update[ "from" ]["userId"] != 0:
                user = self.usernameFromId( update[ "from" ]["userId"] ) #FILL IN
                command = update[ "message"][1:]
        except KeyError as msg:
            traceback.print_exc()
            return None
        return {"user": user, "command": command}

    def execute( self, room, cmd ):
        # valid commands: start
        user = cmd["user"]
        auth = False
        if user in self.authusers:
            auth = True
        cmd = cmd["command"]
        
        if cmd == "start" and room.phase in ['pre','in'] and auth:
            room.talk( self.disclaimer.upper() )
            room.talk( self.startMsg )
            sleep( self.delay )
            room.curInfo()
            start = room.startGame()
            print( "start=", start )
            room.phase = "in"            
            room.gameEvents[ room.phase ].append( room.lastSettings() )
            room.gameEvents[ room.phase ].append( start )
            return
            
        if cmd.startswith( "timers " ) and room.phase == 'pre' and auth:
            try:
                cmd = cmd.split()
                day = int( cmd[1] )
                night = int( cmd[2] )
                room.timers( day, night )
                room.talk( "Day/night set to {}/{}.".format( day, night ) )
            except:
                traceback.print_exc()
            return

        if cmd == "afk" and room.phase in ['pre','in'] and auth:
            try:
                room.afkCheck()
                room.becomeSpectator() 
                room.talk( self.disclaimer )
                room.talk( self.afkMsg )
            except:
                traceback.print_exc()
            return

        if cmd == "boost" and room.phase in ['pre','in'] and auth:
            try:
                room.becomePlayer()
                room.talk( "Boosting..." )
            except:
                traceback.print_exc()
            return

        if cmd == "coffee" and room.phase in['pre','in'] and auth:
            try:
                room.setupCoffee()
                room.talk( "Set up coffee break debug" )
                self.recording = False
            except:
                traceback.print_exc()
            return

        if cmd == "2p" and room.phase in['pre','in'] and auth:
            try:
                room.setup2p()
                room.talk( "Set up 2p debug" )
                self.recording = False
            except:
                traceback.print_exc()
            return

        if cmd == "allstars" and room.phase in ['pre','in'] and auth:
            try:
                room.setupAllStars( self.nextRoomNum )
                room.talk( "Set up All Stars" )
                self.recording = True
            except:
                traceback.print_exc()
            return

        if cmd == "new" and auth:
            try:
                if room.phase == "post":
                    room.talk( "Hosting new room. URL will be provided. Leaving room..." )
                    newRoom = self.hostNewRoom( referer=room.url() )
                    self.recording = True
                    
                    #newRoom.setupAllStars()
                    #print( "hi" )
                    
                
                    self.rooms.remove( room )
                    #room.talk( "Leaving room..." )
                    del room                    
            except:
                traceback.print_exc()
            return

        if cmd == "rooms" and auth:
            try:
                roomStr = "Currently active rooms: "
                for item in self.rooms:
                    roomStr += item.url()+", "
                room.talk( roomStr[:-2] )
            except:
                traceback.print_exc()
            return

        if cmd.startswith( "ban " ) and auth and cmd.split()[1] != "":
            try:
                tmp = cmd.split( None, 2 )[1:]
                toBan = tmp[0]

                
                if len( tmp ) == 1:
                    reason = "-"
                else:
                    reason = tmp[1]
            except:
                traceback.print_exc()
                return
                
            try:
                print( self.blacklist( toBan, reason ) )
                room.talk( toBan + " has been banned. Reason: "+ reason )
            except:
                traceback.print_exc()
                
            return

        if cmd.startswith( "unban " ) and auth and cmd.split()[1] != "":
            toUnBan = cmd.split()[1]
            if toUnBan != "":
                try:
                    print( self.unblacklist( toUnBan ) )
                    room.talk( toUnBan + " has been unbanned." )
                except:
                    traceback.print_exc()
            return

        if cmd == "bans" and auth:
            try:
                room.talk( "Current banlist: "+str( list( self.bans.keys() ) ) )
            except:
                traceback.print_exc()
            return

        if cmd == "clearbans" and auth:
            try:
                self.clearbans()
                room.talk( "Cleared all bans!")
            except:
                traceback.print_exc()
            return

        if cmd.startswith( "kick" ) and room.phase=="pre" and auth:
            toKick = cmd.split()[1]
            if toKick != "": 
                try:
                    rsp = room.kick( self, toKick )
                    if rsp == -1:
                        room.talk( toKick + " not found." )
                    else:
                        print( rsp )
                        room.talk( toKick + " has been kicked." )                       
                    
                except:
                    traceback.print_exc()
            return

        if cmd == "info" and auth:
            try:
                room.talk( "Current users: " + str( room.getUsers() ) )
            except:
                traceback.print_exc()
            return

        if cmd == "list" and auth and room.phase in ["pre" ]:
            try:
                room.list()
                room.talk( "Room was made public." )
            except:
                traceback.print_exc()

        if cmd == "unlist" and auth and room.phase in ["pre" ]:
            try:
                room.unlist()
                room.talk( "Room was made private." )
            except:
                traceback.print_exc()

        if cmd == "exit" and auth and room.phase in ["pre", "post"]:
            #print( "hi" )
            if room.phase == "pre":
                try:
                    room.emptySetup()
                    room.afkCheck()
                    room.becomeSpectator()
                    start = room.startGame()
                    room.phase = "post"
                    
                except:
                    traceback.print_exc()
            try:
                room.talk( "Session ended, goodbye." )
                
            except:
                traceback.print_exc()
                
            self.keepGoing = False

        if cmd == "ping" and auth:
            try:
                room.talk("pong")
            except:
                traceback.print_exc()
        
            
                

# everything below will go in main

def mainLoop( sesh ):
    while sesh.keepGoing:
        for room in sesh.rooms:
            update = room.getNextEvent()
            if DEBUG:
                print( "Processing:", update )
            if update:
                
                room.gameEvents[ room.phase ].append( update )
                cmd = sesh.processPacket( room, update )
                if cmd:
                    sesh.execute( room, cmd )
            #print( "Last event:", room.gameEvents[ room.phase ][-1] )
        # ignoring messages from host in these logs?
                           
    

def main():
    # create new session
    sesh = Session( "credentials.txt", "dicts.txt" )
    print( sesh )
    sesh.hostNewRoom()    
    
    #sesh.rooms[0].talk( "Hi" )
    #sleep( 15 )
    #print( "doing afk check" )
    #print( "Doing afk check returned:", sesh.rooms[0].afkCheck() )
    #print( "Room info:", sesh.rooms[0].curInfo() )
    #print( "About user:", sesh.userInfo( 1000 ) )
    #sesh.rooms[0].setupAllStars()
   
    #print( "Changing setup returned:", sesh.rooms[0].setup( "75a1b93a1" ) )
    #print( "Changing timers returned:", sesh.rooms[0].timers( 1, 1 ) )
    #print( "Changing timers returned:", sesh.rooms[0].scaleTimer( True ) )
    #print( "Setting vote lock returned:", sesh.rooms[0].disableVoteLock( True ) )
    #print( "Room info again:", sesh.rooms[0].curInfo() )
    #print( "Becoming spec returns:", sesh.rooms[0].becomeSpectator() )

    #print( "Room info from session: ", sesh.roomInfo( sesh.rooms[0].roomId ) )
    #sleep( 10 )
    #print( "Starting game" )
    #sesh.rooms[0].startGame()
    mainLoop( sesh )
    #while True:
        #sleep( 1 )
    #    print( "Latest event:", sesh.rooms[0].getNextEvent() )
        
    #sesh.rooms[0].startGame()
    #sleep( 315 )
    #print( "Room info again:", sesh.rooms[0].curInfo() )
    print( "Goodbye" )

main()
        
        
